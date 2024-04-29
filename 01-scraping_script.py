import requests
from bs4 import BeautifulSoup
import openpyxl

workbook = openpyxl.Workbook()
active_sheet = workbook.active

url = requests.get("https://www.infotec.com.pe/3-laptops-y-notebooks")
soup = BeautifulSoup(url.content, 'html.parser')

names = []

elements_a = soup.find_all('h2',class_='h3 product-title')

for element_h2 in elements_a:
    element_a = element_h2.find('a')
    if element_a:
        names.append(element_a.text.strip())

prices = []

elements_b = soup.find_all('div', class_='product-price-and-shipping')

for element_div in elements_b:
    element_span = element_div.find('span')
    if element_span:
        prices.append(element_span.text.strip())

for i in range(len(names)):
    active_sheet.cell(row=i+1, column=1, value=names[i])
    active_sheet.cell(row=i+1, column=2, value=prices[i])

workbook.save("prices.xlsx")